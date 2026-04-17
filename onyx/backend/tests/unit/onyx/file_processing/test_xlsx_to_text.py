import io
from typing import cast
from unittest.mock import MagicMock

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from onyx.file_processing.extract_file_text import _clean_worksheet_matrix
from onyx.file_processing.extract_file_text import _worksheet_to_matrix
from onyx.file_processing.extract_file_text import xlsx_sheet_extraction
from onyx.file_processing.extract_file_text import xlsx_to_text


def _make_xlsx(sheets: dict[str, list[list[str]]]) -> io.BytesIO:
    """Create an in-memory xlsx file from a dict of sheet_name -> matrix of strings."""
    wb = openpyxl.Workbook()
    if wb.active is not None:
        wb.remove(cast(Worksheet, wb.active))
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestXlsxToText:
    def test_single_sheet_basic(self) -> None:
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ["Name", "Age"],
                    ["Alice", "30"],
                    ["Bob", "25"],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        lines = [line for line in result.strip().split("\n") if line.strip()]
        assert len(lines) == 3
        assert "Name" in lines[0]
        assert "Age" in lines[0]
        assert "Alice" in lines[1]
        assert "30" in lines[1]
        assert "Bob" in lines[2]

    def test_multiple_sheets_separated(self) -> None:
        xlsx = _make_xlsx(
            {
                "Sheet1": [["a", "b"]],
                "Sheet2": [["c", "d"]],
            }
        )
        result = xlsx_to_text(xlsx)
        # TEXT_SECTION_SEPARATOR is "\n\n"
        assert "\n\n" in result
        parts = result.split("\n\n")
        assert any("a" in p for p in parts)
        assert any("c" in p for p in parts)

    def test_empty_cells(self) -> None:
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ["a", "", "b"],
                    ["", "c", ""],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        lines = [line for line in result.strip().split("\n") if line.strip()]
        assert len(lines) == 2

    def test_commas_in_cells_are_quoted(self) -> None:
        """Cells containing commas should be quoted in CSV output."""
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ["hello, world", "normal"],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        assert '"hello, world"' in result

    def test_empty_workbook(self) -> None:
        xlsx = _make_xlsx({"Sheet1": []})
        result = xlsx_to_text(xlsx)
        assert result.strip() == ""

    def test_long_empty_row_run_capped(self) -> None:
        """Runs of >2 empty rows should be capped to 2."""
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ["header"],
                    [""],
                    [""],
                    [""],
                    [""],
                    ["data"],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        lines = [line for line in result.strip().split("\n") if line.strip()]
        # 4 empty rows capped to 2, so: header + 2 empty + data = 4 lines
        assert len(lines) == 4
        assert "header" in lines[0]
        assert "data" in lines[-1]

    def test_long_empty_col_run_capped(self) -> None:
        """Runs of >2 empty columns should be capped to 2."""
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ["a", "", "", "", "b"],
                    ["c", "", "", "", "d"],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        lines = [line for line in result.strip().split("\n") if line.strip()]
        assert len(lines) == 2
        # Each row should have 4 fields (a + 2 empty + b), not 5
        # csv format: a,,,b (3 commas = 4 fields)
        first_line = lines[0].strip()
        # Count commas to verify column reduction
        assert first_line.count(",") == 3

    def test_short_empty_runs_kept(self) -> None:
        """Runs of <=2 empty rows/cols should be preserved."""
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ["a", "b"],
                    ["", ""],
                    ["", ""],
                    ["c", "d"],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        lines = [line for line in result.strip().split("\n") if line.strip()]
        # All 4 rows preserved (2 empty rows <= threshold)
        assert len(lines) == 4

    def test_bad_zip_file_returns_empty(self) -> None:
        bad_file = io.BytesIO(b"not a zip file")
        result = xlsx_to_text(bad_file, file_name="test.xlsx")
        assert result == ""

    def test_bad_zip_tilde_file_returns_empty(self) -> None:
        bad_file = io.BytesIO(b"not a zip file")
        result = xlsx_to_text(bad_file, file_name="~$temp.xlsx")
        assert result == ""

    def test_large_sparse_sheet(self) -> None:
        """A sheet with data, a big empty gap, and more data — gap is capped to 2."""
        rows: list[list[str]] = [["row1_data"]]
        rows.extend([[""] for _ in range(10)])
        rows.append(["row2_data"])
        xlsx = _make_xlsx({"Sheet1": rows})
        result = xlsx_to_text(xlsx)
        lines = [line for line in result.strip().split("\n") if line.strip()]
        # 10 empty rows capped to 2: row1_data + 2 empty + row2_data = 4
        assert len(lines) == 4
        assert "row1_data" in lines[0]
        assert "row2_data" in lines[-1]

    def test_quotes_in_cells(self) -> None:
        """Cells containing quotes should be properly escaped."""
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ['say "hello"', "normal"],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        # csv.writer escapes quotes by doubling them
        assert '""hello""' in result

    def test_each_row_is_separate_line(self) -> None:
        """Each row should produce its own line (regression for writerow vs writerows)."""
        xlsx = _make_xlsx(
            {
                "Sheet1": [
                    ["r1c1", "r1c2"],
                    ["r2c1", "r2c2"],
                    ["r3c1", "r3c2"],
                ]
            }
        )
        result = xlsx_to_text(xlsx)
        lines = [line for line in result.strip().split("\n") if line.strip()]
        assert len(lines) == 3
        assert "r1c1" in lines[0] and "r1c2" in lines[0]
        assert "r2c1" in lines[1] and "r2c2" in lines[1]
        assert "r3c1" in lines[2] and "r3c2" in lines[2]


class TestWorksheetToMatrixJaggedRows:
    """openpyxl read_only mode can yield rows of differing widths when
    trailing cells are empty. The matrix must be padded to a rectangle
    so downstream column cleanup can index safely."""

    def test_pads_shorter_trailing_rows(self) -> None:
        ws = MagicMock()
        ws.iter_rows.return_value = iter(
            [
                ("A", "B", "C"),
                ("X", "Y"),
                ("P",),
            ]
        )
        matrix = _worksheet_to_matrix(ws)
        assert matrix == [["A", "B", "C"], ["X", "Y", ""], ["P", "", ""]]

    def test_pads_when_first_row_is_shorter(self) -> None:
        ws = MagicMock()
        ws.iter_rows.return_value = iter(
            [
                ("A",),
                ("X", "Y", "Z"),
            ]
        )
        matrix = _worksheet_to_matrix(ws)
        assert matrix == [["A", "", ""], ["X", "Y", "Z"]]

    def test_clean_worksheet_matrix_no_index_error_on_jagged_rows(self) -> None:
        """Regression: previously raised IndexError when a later row was
        shorter than the first row and the out-of-range column on the
        first row was empty (so the short-circuit in `all()` did not
        save us)."""
        ws = MagicMock()
        ws.iter_rows.return_value = iter(
            [
                ("A", "", "", "B"),
                ("X", "Y"),
            ]
        )
        matrix = _worksheet_to_matrix(ws)
        # Must not raise.
        cleaned = _clean_worksheet_matrix(matrix)
        assert cleaned == [["A", "", "", "B"], ["X", "Y", "", ""]]


class TestXlsxSheetExtraction:
    def test_one_tuple_per_sheet(self) -> None:
        xlsx = _make_xlsx(
            {
                "Revenue": [["Month", "Amount"], ["Jan", "100"]],
                "Expenses": [["Category", "Cost"], ["Rent", "500"]],
            }
        )
        sheets = xlsx_sheet_extraction(xlsx)
        assert len(sheets) == 2
        # Order preserved from workbook sheet order
        titles = [title for _csv, title in sheets]
        assert titles == ["Revenue", "Expenses"]
        # Content present in the right tuple
        revenue_csv, _ = sheets[0]
        expenses_csv, _ = sheets[1]
        assert "Month" in revenue_csv
        assert "Jan" in revenue_csv
        assert "Category" in expenses_csv
        assert "Rent" in expenses_csv

    def test_tuple_structure_is_csv_text_then_title(self) -> None:
        """The tuple order is (csv_text, sheet_title) — pin it so callers
        that unpack positionally don't silently break."""
        xlsx = _make_xlsx({"MySheet": [["a", "b"]]})
        sheets = xlsx_sheet_extraction(xlsx)
        assert len(sheets) == 1
        csv_text, title = sheets[0]
        assert title == "MySheet"
        assert "a" in csv_text
        assert "b" in csv_text

    def test_empty_sheet_is_skipped(self) -> None:
        """A sheet whose CSV output is empty/whitespace-only should NOT
        appear in the result — the `if csv_text.strip():` guard filters
        it out."""
        xlsx = _make_xlsx(
            {
                "Data": [["a", "b"]],
                "Empty": [],
            }
        )
        sheets = xlsx_sheet_extraction(xlsx)
        assert len(sheets) == 1
        assert sheets[0][1] == "Data"

    def test_empty_workbook_returns_empty_list(self) -> None:
        """All sheets empty → empty list (not a list of empty tuples)."""
        xlsx = _make_xlsx({"Sheet1": [], "Sheet2": []})
        sheets = xlsx_sheet_extraction(xlsx)
        assert sheets == []

    def test_single_sheet(self) -> None:
        xlsx = _make_xlsx({"Only": [["x", "y"], ["1", "2"]]})
        sheets = xlsx_sheet_extraction(xlsx)
        assert len(sheets) == 1
        csv_text, title = sheets[0]
        assert title == "Only"
        assert "x" in csv_text
        assert "1" in csv_text

    def test_bad_zip_returns_empty_list(self) -> None:
        bad_file = io.BytesIO(b"not a zip file")
        sheets = xlsx_sheet_extraction(bad_file, file_name="test.xlsx")
        assert sheets == []

    def test_bad_zip_tilde_file_returns_empty_list(self) -> None:
        """`~$`-prefixed files are Excel lock files; failure should log
        at debug (not warning) and still return []."""
        bad_file = io.BytesIO(b"not a zip file")
        sheets = xlsx_sheet_extraction(bad_file, file_name="~$temp.xlsx")
        assert sheets == []

    def test_csv_content_matches_xlsx_to_text_per_sheet(self) -> None:
        """For a single-sheet workbook, xlsx_to_text output should equal
        the csv_text from xlsx_sheet_extraction — they share the same
        per-sheet CSV-ification logic."""
        single_sheet_data = [["Name", "Age"], ["Alice", "30"]]
        expected_text = xlsx_to_text(_make_xlsx({"People": single_sheet_data}))

        sheets = xlsx_sheet_extraction(_make_xlsx({"People": single_sheet_data}))
        assert len(sheets) == 1
        csv_text, title = sheets[0]
        assert title == "People"
        assert csv_text.strip() == expected_text.strip()

    def test_commas_in_cells_are_quoted(self) -> None:
        xlsx = _make_xlsx({"S1": [["hello, world", "normal"]]})
        sheets = xlsx_sheet_extraction(xlsx)
        assert len(sheets) == 1
        csv_text, _ = sheets[0]
        assert '"hello, world"' in csv_text

    def test_long_empty_row_run_capped_within_sheet(self) -> None:
        """The matrix cleanup applies per-sheet: >2 empty rows collapse
        to 2, which keeps the sheet non-empty and it still appears in
        the result."""
        xlsx = _make_xlsx(
            {
                "S1": [
                    ["header"],
                    [""],
                    [""],
                    [""],
                    [""],
                    ["data"],
                ]
            }
        )
        sheets = xlsx_sheet_extraction(xlsx)
        assert len(sheets) == 1
        csv_text, _ = sheets[0]
        lines = csv_text.strip().split("\n")
        # header + 2 empty (capped) + data = 4 lines
        assert len(lines) == 4
        assert "header" in lines[0]
        assert "data" in lines[-1]

    def test_sheet_title_with_special_chars_preserved(self) -> None:
        """Spaces, punctuation, unicode in sheet titles are preserved
        verbatim — the title is used as a link anchor downstream."""
        xlsx = _make_xlsx(
            {
                "Q1 Revenue (USD)": [["a", "b"]],
                "Données": [["c", "d"]],
            }
        )
        sheets = xlsx_sheet_extraction(xlsx)
        titles = [title for _csv, title in sheets]
        assert "Q1 Revenue (USD)" in titles
        assert "Données" in titles
