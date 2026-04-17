import csv
import gc
import io
import json
import os
import re
import zipfile
from collections.abc import Callable
from collections.abc import Iterator
from collections.abc import Sequence
from email.parser import Parser as EmailParser
from io import BytesIO
from pathlib import Path
from typing import Any
from typing import IO
from typing import NamedTuple
from typing import Optional
from typing import TYPE_CHECKING
from zipfile import BadZipFile

import chardet
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from onyx.configs.constants import ONYX_METADATA_FILENAME
from onyx.configs.llm_configs import get_image_extraction_and_analysis_enabled
from onyx.file_processing.file_types import OnyxFileExtensions
from onyx.file_processing.file_types import OnyxMimeTypes
from onyx.file_processing.file_types import PRESENTATION_MIME_TYPE
from onyx.file_processing.file_types import WORD_PROCESSING_MIME_TYPE
from onyx.file_processing.html_utils import parse_html_page_basic
from onyx.file_processing.unstructured import get_unstructured_api_key
from onyx.file_processing.unstructured import unstructured_to_text
from onyx.utils.logger import setup_logger

if TYPE_CHECKING:
    from markitdown import MarkItDown
logger = setup_logger()

TEXT_SECTION_SEPARATOR = "\n\n"

_MARKITDOWN_CONVERTER: Optional["MarkItDown"] = None

KNOWN_OPENPYXL_BUGS = [
    "Value must be either numerical or a string containing a wildcard",
    "File contains no valid workbook part",
    "Unable to read workbook: could not read stylesheet from None",
    "Colors must be aRGB hex values",
]


def get_markitdown_converter() -> "MarkItDown":
    global _MARKITDOWN_CONVERTER

    if _MARKITDOWN_CONVERTER is None:
        from markitdown import MarkItDown

        # Patch this function to effectively no-op because we were seeing this
        # module take an inordinate amount of time to convert charts to markdown,
        # making some powerpoint files with many or complicated charts nearly
        # unindexable.
        from markitdown.converters._pptx_converter import PptxConverter

        setattr(
            PptxConverter,
            "_convert_chart_to_markdown",
            lambda self, chart: "\n\n[chart omitted]\n\n",  # noqa: ARG005
        )
        _MARKITDOWN_CONVERTER = MarkItDown(enable_plugins=False)
    return _MARKITDOWN_CONVERTER


def get_file_ext(file_path_or_name: str | Path) -> str:
    _, extension = os.path.splitext(file_path_or_name)
    return extension.lower()


def is_text_file(file: IO[bytes]) -> bool:
    """
    checks if the first 1024 bytes only contain printable or whitespace characters
    if it does, then we say it's a plaintext file
    """
    raw_data = file.read(1024)
    file.seek(0)
    text_chars = bytearray({7, 8, 9, 10, 12, 13, 27} | set(range(0x20, 0x100)) - {0x7F})
    return all(c in text_chars for c in raw_data)


def detect_encoding(file: IO[bytes]) -> str:
    raw_data = file.read(50000)
    file.seek(0)
    encoding = chardet.detect(raw_data)["encoding"] or "utf-8"
    return encoding


def is_macos_resource_fork_file(file_name: str) -> bool:
    return os.path.basename(file_name).startswith("._") and file_name.startswith(
        "__MACOSX"
    )


def to_bytesio(stream: IO[bytes]) -> BytesIO:
    if isinstance(stream, BytesIO):
        return stream
    data = stream.read()  # consumes the stream!
    return BytesIO(data)


def load_files_from_zip(
    zip_file_io: IO,
    ignore_macos_resource_fork_files: bool = True,
    ignore_dirs: bool = True,
) -> Iterator[tuple[zipfile.ZipInfo, IO[Any]]]:
    """
    Iterates through files in a zip archive, yielding (ZipInfo, file handle) pairs.
    """
    with zipfile.ZipFile(zip_file_io, "r") as zip_file:
        for file_info in zip_file.infolist():
            if ignore_dirs and file_info.is_dir():
                continue

            if (
                ignore_macos_resource_fork_files
                and is_macos_resource_fork_file(file_info.filename)
            ) or file_info.filename == ONYX_METADATA_FILENAME:
                continue

            with zip_file.open(file_info.filename, "r") as subfile:
                # Try to match by exact filename first
                yield file_info, subfile


def _extract_onyx_metadata(line: str) -> dict | None:
    """
    Example: first line has:
        <!-- ONYX_METADATA={"title": "..."} -->
      or
        #ONYX_METADATA={"title":"..."}
    """
    html_comment_pattern = r"<!--\s*ONYX_METADATA=\{(.*?)\}\s*-->"
    hashtag_pattern = r"#ONYX_METADATA=\{(.*?)\}"

    html_comment_match = re.search(html_comment_pattern, line)
    hashtag_match = re.search(hashtag_pattern, line)

    if html_comment_match:
        json_str = html_comment_match.group(1)
    elif hashtag_match:
        json_str = hashtag_match.group(1)
    else:
        return None

    try:
        return json.loads("{" + json_str + "}")
    except json.JSONDecodeError:
        return None


def read_text_file(
    file: IO,
    encoding: str = "utf-8",
    errors: str = "replace",
    ignore_onyx_metadata: bool = True,
) -> tuple[str, dict]:
    """
    For plain text files. Optionally extracts Onyx metadata from the first line.
    """
    metadata = {}
    file_content_raw = ""
    for ind, line in enumerate(file):
        # decode
        try:
            line = line.decode(encoding) if isinstance(line, bytes) else line
        except UnicodeDecodeError:
            line = (
                line.decode(encoding, errors=errors)
                if isinstance(line, bytes)
                else line
            )

        # optionally parse metadata in the first line
        if ind == 0 and not ignore_onyx_metadata:
            potential_meta = _extract_onyx_metadata(line)
            if potential_meta is not None:
                metadata = potential_meta
                continue

        file_content_raw += line

    return file_content_raw, metadata


def pdf_to_text(file: IO[Any], pdf_pass: str | None = None) -> str:
    """
    Extract text from a PDF. For embedded images, a more complex approach is needed.
    This is a minimal approach returning text only.
    """
    text, _, _ = read_pdf_file(file, pdf_pass)
    return text


def read_pdf_file(
    file: IO[Any],
    pdf_pass: str | None = None,
    extract_images: bool = False,
    image_callback: Callable[[bytes, str], None] | None = None,
) -> tuple[str, dict[str, Any], Sequence[tuple[bytes, str]]]:
    """
    Returns the text, basic PDF metadata, and optionally extracted images.
    """
    from pypdf import PdfReader
    from pypdf.errors import PdfStreamError

    metadata: dict[str, Any] = {}
    extracted_images: list[tuple[bytes, str]] = []
    try:
        pdf_reader = PdfReader(file)

        if pdf_reader.is_encrypted:
            # Try the explicit password first, then fall back to an empty
            # string.  Owner-password-only PDFs (permission restrictions but
            # no open password) decrypt successfully with "".
            # See https://github.com/onyx-dot-app/onyx/issues/9754
            passwords = [p for p in [pdf_pass, ""] if p is not None]
            decrypt_success = False
            for pw in passwords:
                try:
                    if pdf_reader.decrypt(pw) != 0:
                        decrypt_success = True
                        break
                except Exception:
                    pass

            if not decrypt_success:
                logger.error(
                    "Encrypted PDF could not be decrypted, returning empty text."
                )
                return "", metadata, []

        # Basic PDF metadata
        if pdf_reader.metadata is not None:
            for key, value in pdf_reader.metadata.items():
                clean_key = key.lstrip("/")
                if isinstance(value, str) and value.strip():
                    metadata[clean_key] = value
                elif isinstance(value, list) and all(
                    isinstance(item, str) for item in value
                ):
                    metadata[clean_key] = ", ".join(value)

        text = TEXT_SECTION_SEPARATOR.join(
            page.extract_text() for page in pdf_reader.pages
        )

        if extract_images:
            for page_num, page in enumerate(pdf_reader.pages):
                for image_file_object in page.images:
                    image = Image.open(io.BytesIO(image_file_object.data))
                    img_byte_arr = io.BytesIO()
                    image.save(img_byte_arr, format=image.format)
                    img_bytes = img_byte_arr.getvalue()

                    image_format = image.format.lower() if image.format else "png"
                    image_name = f"page_{page_num + 1}_image_{image_file_object.name}.{image_format}"
                    if image_callback is not None:
                        # Stream image out immediately
                        image_callback(img_bytes, image_name)
                    else:
                        extracted_images.append((img_bytes, image_name))

        return text, metadata, extracted_images

    except PdfStreamError:
        logger.exception("Invalid PDF file")
    except Exception:
        logger.exception("Failed to read PDF")

    return "", metadata, []


def extract_docx_images(docx_bytes: IO[Any]) -> Iterator[tuple[bytes, str]]:
    """
    Given the bytes of a docx file, extract all the images.
    Returns a list of tuples (image_bytes, image_name).
    """
    try:
        with zipfile.ZipFile(docx_bytes) as z:
            for name in z.namelist():
                if name.startswith("word/media/"):
                    yield (z.read(name), name.split("/")[-1])
    except Exception:
        logger.exception("Failed to extract all docx images")


def read_docx_file(
    file: IO[Any],
    file_name: str = "",
    extract_images: bool = False,
    image_callback: Callable[[bytes, str], None] | None = None,
) -> tuple[str, Sequence[tuple[bytes, str]]]:
    """
    Extract text from a docx.
    Return (text_content, list_of_images).

    The caller can choose to provide a callback to handle images with the intent
    of avoiding materializing the list of images in memory.
    The images list returned is empty in this case.
    """
    md = get_markitdown_converter()
    from markitdown import (
        StreamInfo,
        FileConversionException,
        UnsupportedFormatException,
    )

    try:
        doc = md.convert(
            to_bytesio(file), stream_info=StreamInfo(mimetype=WORD_PROCESSING_MIME_TYPE)
        )
    except (
        BadZipFile,
        ValueError,
        FileConversionException,
        UnsupportedFormatException,
    ) as e:
        logger.warning(
            f"Failed to extract docx {file_name or 'docx file'}: {e}. Attempting to read as text file."
        )

        # May be an invalid docx, but still a valid text file
        file.seek(0)
        encoding = detect_encoding(file)
        text_content_raw, _ = read_text_file(
            file, encoding=encoding, ignore_onyx_metadata=False
        )
        return text_content_raw or "", []

    file.seek(0)

    if extract_images:
        if image_callback is None:
            return doc.markdown, list(extract_docx_images(to_bytesio(file)))
        # If a callback is provided, iterate and stream images without accumulating
        try:
            for img_file_bytes, img_file_name in extract_docx_images(to_bytesio(file)):
                image_callback(img_file_bytes, img_file_name)
        except Exception:
            logger.exception("Failed to stream docx images")
    return doc.markdown, []


def pptx_to_text(file: IO[Any], file_name: str = "") -> str:
    md = get_markitdown_converter()
    from markitdown import (
        StreamInfo,
        FileConversionException,
        UnsupportedFormatException,
    )

    stream_info = StreamInfo(
        mimetype=PRESENTATION_MIME_TYPE, filename=file_name or None, extension=".pptx"
    )
    try:
        presentation = md.convert(to_bytesio(file), stream_info=stream_info)
    except (
        BadZipFile,
        ValueError,
        FileConversionException,
        UnsupportedFormatException,
    ) as e:
        error_str = f"Failed to extract text from {file_name or 'pptx file'}: {e}"
        logger.warning(error_str)
        return ""
    return presentation.markdown


def _worksheet_to_matrix(
    worksheet: Worksheet,
) -> list[list[str]]:
    """
    Converts a singular worksheet to a matrix of values.

    Rows are padded to a uniform width. In openpyxl's read_only mode,
    iter_rows can yield rows of differing lengths (trailing empty cells
    are sometimes omitted), and downstream column cleanup assumes a
    rectangular matrix.
    """
    rows: list[list[str]] = []
    max_len = 0
    for worksheet_row in worksheet.iter_rows(min_row=1, values_only=True):
        row = ["" if cell is None else str(cell) for cell in worksheet_row]
        if len(row) > max_len:
            max_len = len(row)
        rows.append(row)

    for row in rows:
        if len(row) < max_len:
            row.extend([""] * (max_len - len(row)))

    return rows


def _clean_worksheet_matrix(matrix: list[list[str]]) -> list[list[str]]:
    """
    Cleans a worksheet matrix by removing rows if there are N consecutive empty
    rows and removing cols if there are M consecutive empty columns
    """
    MAX_EMPTY_ROWS = 2  # Runs longer than this are capped to max_empty; shorter runs are preserved as-is
    MAX_EMPTY_COLS = 2

    # Row cleanup
    matrix = _remove_empty_runs(matrix, max_empty=MAX_EMPTY_ROWS)

    if not matrix:
        return matrix

    # Column cleanup — determine which columns to keep without transposing.
    num_cols = len(matrix[0])
    keep_cols = _columns_to_keep(matrix, num_cols, max_empty=MAX_EMPTY_COLS)
    if len(keep_cols) < num_cols:
        matrix = [[row[c] for c in keep_cols] for row in matrix]

    return matrix


def _columns_to_keep(
    matrix: list[list[str]], num_cols: int, max_empty: int
) -> list[int]:
    """Return the indices of columns to keep after removing empty-column runs.

    Uses the same logic as ``_remove_empty_runs`` but operates on column
    indices so no transpose is needed.
    """
    kept: list[int] = []
    empty_buffer: list[int] = []

    for col_idx in range(num_cols):
        col_is_empty = all(not row[col_idx] for row in matrix)
        if col_is_empty:
            empty_buffer.append(col_idx)
        else:
            kept.extend(empty_buffer[:max_empty])
            kept.append(col_idx)
            empty_buffer = []

    return kept


def _remove_empty_runs(
    rows: list[list[str]],
    max_empty: int,
) -> list[list[str]]:
    """Removes entire runs of empty rows when the run length exceeds max_empty.

    Leading empty runs are capped to max_empty, just like interior runs.
    Trailing empty rows are always dropped since there is no subsequent
    non-empty row to flush them.
    """
    result: list[list[str]] = []
    empty_buffer: list[list[str]] = []

    for row in rows:
        # Check if empty
        if not any(row):
            if len(empty_buffer) < max_empty:
                empty_buffer.append(row)
        else:
            # Add upto max empty rows onto the result - that's what we allow
            result.extend(empty_buffer[:max_empty])
            # Add the new non-empty row
            result.append(row)
            empty_buffer = []

    return result


def xlsx_sheet_extraction(file: IO[Any], file_name: str = "") -> list[tuple[str, str]]:
    """
    Converts each sheet in the excel file to a csv condensed string.
    Returns a string and the worksheet title for each worksheet

    Returns a list of (csv_text, sheet)
    """
    try:
        workbook = openpyxl.load_workbook(file, read_only=True)
    except BadZipFile as e:
        error_str = f"Failed to extract text from {file_name or 'xlsx file'}: {e}"
        if file_name.startswith("~"):
            logger.debug(error_str + " (this is expected for files with ~)")
        else:
            logger.warning(error_str)
        return []
    except Exception as e:
        if any(s in str(e) for s in KNOWN_OPENPYXL_BUGS):
            logger.error(
                f"Failed to extract text from {file_name or 'xlsx file'}. This happens due to a bug in openpyxl. {e}"
            )
            return []
        raise

    sheets: list[tuple[str, str]] = []
    for sheet in workbook.worksheets:
        sheet_matrix = _clean_worksheet_matrix(_worksheet_to_matrix(sheet))
        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\n")
        writer.writerows(sheet_matrix)
        csv_text = buf.getvalue().rstrip("\n")
        if csv_text.strip():
            sheets.append((csv_text, sheet.title))
    return sheets


def xlsx_to_text(file: IO[Any], file_name: str = "") -> str:
    sheets = xlsx_sheet_extraction(file, file_name)
    return TEXT_SECTION_SEPARATOR.join(csv_text for csv_text, _title in sheets)


def eml_to_text(file: IO[Any]) -> str:
    encoding = detect_encoding(file)
    text_file = io.TextIOWrapper(file, encoding=encoding)
    parser = EmailParser()
    try:
        message = parser.parse(text_file)
    finally:
        try:
            # Keep underlying upload handle open for downstream consumers.
            raw_file = text_file.detach()
        except Exception as detach_error:
            logger.warning(
                f"Failed to detach TextIOWrapper for EML upload, using original file: {detach_error}"
            )
            raw_file = file
        try:
            raw_file.seek(0)
        except Exception:
            pass

    text_content = []
    for part in message.walk():
        if part.get_content_type().startswith("text/plain"):
            payload = part.get_payload()
            if isinstance(payload, str):
                text_content.append(payload)
            elif isinstance(payload, list):
                text_content.extend(item for item in payload if isinstance(item, str))
            else:
                logger.warning(f"Unexpected payload type: {type(payload)}")
    return TEXT_SECTION_SEPARATOR.join(text_content)


def epub_to_text(file: IO[Any]) -> str:
    with zipfile.ZipFile(file) as epub:
        text_content = []
        for item in epub.infolist():
            if item.filename.endswith(".xhtml") or item.filename.endswith(".html"):
                with epub.open(item) as html_file:
                    text_content.append(parse_html_page_basic(html_file))
        return TEXT_SECTION_SEPARATOR.join(text_content)


def file_io_to_text(file: IO[Any]) -> str:
    encoding = detect_encoding(file)
    file_content, _ = read_text_file(file, encoding=encoding)
    return file_content


def extract_file_text(
    file: IO[Any],
    file_name: str,
    break_on_unprocessable: bool = True,
    extension: str | None = None,
) -> str:
    """
    Legacy function that returns *only text*, ignoring embedded images.
    For backward-compatibility in code that only wants text.

    NOTE: Ignoring seems to be defined as returning an empty string for files it can't
    handle (such as images).
    """
    extension_to_function: dict[str, Callable[[IO[Any]], str]] = {
        ".pdf": pdf_to_text,
        ".docx": lambda f: read_docx_file(f, file_name)[0],  # no images
        ".pptx": lambda f: pptx_to_text(f, file_name),
        ".xlsx": lambda f: xlsx_to_text(f, file_name),
        ".eml": eml_to_text,
        ".epub": epub_to_text,
        ".html": parse_html_page_basic,
    }

    try:
        if get_unstructured_api_key():
            try:
                return unstructured_to_text(file, file_name)
            except Exception as unstructured_error:
                logger.error(
                    f"Failed to process with Unstructured: {str(unstructured_error)}. Falling back to normal processing."
                )
        if extension is None:
            extension = get_file_ext(file_name)

        if extension in OnyxFileExtensions.TEXT_AND_DOCUMENT_EXTENSIONS:
            func = extension_to_function.get(extension, file_io_to_text)
            file.seek(0)
            return func(file)

        # If unknown extension, maybe it's a text file
        file.seek(0)
        if is_text_file(file):
            return file_io_to_text(file)

        raise ValueError("Unknown file extension or not recognized as text data")

    except Exception as e:
        if break_on_unprocessable:
            raise RuntimeError(
                f"Failed to process file {file_name or 'Unknown'}: {str(e)}"
            ) from e
        logger.warning(f"Failed to process file {file_name or 'Unknown'}: {str(e)}")
        return ""


class ExtractionResult(NamedTuple):
    """Structured result from text and image extraction from various file types."""

    text_content: str
    embedded_images: Sequence[tuple[bytes, str]]
    metadata: dict[str, Any]


def extract_result_from_text_file(file: IO[Any]) -> ExtractionResult:
    encoding = detect_encoding(file)
    text_content_raw, file_metadata = read_text_file(
        file, encoding=encoding, ignore_onyx_metadata=False
    )
    return ExtractionResult(
        text_content=text_content_raw,
        embedded_images=[],
        metadata=file_metadata,
    )


def extract_text_and_images(
    file: IO[Any],
    file_name: str,
    pdf_pass: str | None = None,
    content_type: str | None = None,
    image_callback: Callable[[bytes, str], None] | None = None,
) -> ExtractionResult:
    """
    Primary new function for the updated connector.
    Returns structured extraction result with text content, embedded images, and metadata.

    Args:
        file: File-like object to extract content from.
        file_name: Name of the file (used to determine extension/type).
        pdf_pass: Optional password for encrypted PDFs.
        content_type: Optional MIME type override for the file.
        image_callback: Optional callback for streaming image extraction. When provided,
            embedded images are passed to this callback one at a time as (bytes, filename)
            instead of being accumulated in the returned ExtractionResult.embedded_images
            list. This is a memory optimization for large documents with many images -
            the caller can process/store each image immediately rather than holding all
            images in memory. When using a callback, ExtractionResult.embedded_images
            will be an empty list.

    Returns:
        ExtractionResult containing text_content, embedded_images (empty if callback used),
        and metadata extracted from the file.
    """
    res = _extract_text_and_images(
        file, file_name, pdf_pass, content_type, image_callback
    )
    # Clean up any temporary objects and force garbage collection
    unreachable = gc.collect()
    logger.info(f"Unreachable objects: {unreachable}")

    return res


def _extract_text_and_images(
    file: IO[Any],
    file_name: str,
    pdf_pass: str | None = None,
    content_type: str | None = None,
    image_callback: Callable[[bytes, str], None] | None = None,
) -> ExtractionResult:
    file.seek(0)

    if get_unstructured_api_key():
        try:
            text_content = unstructured_to_text(file, file_name)
            return ExtractionResult(
                text_content=text_content, embedded_images=[], metadata={}
            )
        except Exception as e:
            logger.error(
                f"Failed to process with Unstructured: {str(e)}. Falling back to normal processing."
            )
            file.seek(0)  # Reset file pointer just in case

    # When we upload a document via a connector or MyDocuments, we extract and store the content of files
    # with content types in UploadMimeTypes.DOCUMENT_MIME_TYPES as plain text files.
    # As a result, the file name extension may differ from the original content type.
    # We process files with a plain text content type first to handle this scenario.
    if content_type in OnyxMimeTypes.TEXT_MIME_TYPES:
        return extract_result_from_text_file(file)

    # Default processing
    try:
        extension = get_file_ext(file_name)
        # docx example for embedded images
        if extension == ".docx":
            text_content, images = read_docx_file(
                file, file_name, extract_images=True, image_callback=image_callback
            )
            return ExtractionResult(
                text_content=text_content, embedded_images=images, metadata={}
            )

        # PDF example: we do not show complicated PDF image extraction here
        # so we simply extract text for now and skip images.
        if extension == ".pdf":
            text_content, pdf_metadata, images = read_pdf_file(
                file,
                pdf_pass,
                extract_images=get_image_extraction_and_analysis_enabled(),
                image_callback=image_callback,
            )
            return ExtractionResult(
                text_content=text_content, embedded_images=images, metadata=pdf_metadata
            )

        # For PPTX, XLSX, EML, etc., we do not show embedded image logic here.
        # You can do something similar to docx if needed.
        if extension == ".pptx":
            return ExtractionResult(
                text_content=pptx_to_text(file, file_name=file_name),
                embedded_images=[],
                metadata={},
            )

        if extension == ".xlsx":
            return ExtractionResult(
                text_content=xlsx_to_text(file, file_name=file_name),
                embedded_images=[],
                metadata={},
            )

        if extension == ".eml":
            return ExtractionResult(
                text_content=eml_to_text(file), embedded_images=[], metadata={}
            )

        if extension == ".epub":
            return ExtractionResult(
                text_content=epub_to_text(file), embedded_images=[], metadata={}
            )

        if extension == ".html":
            return ExtractionResult(
                text_content=parse_html_page_basic(file),
                embedded_images=[],
                metadata={},
            )

        # If we reach here and it's a recognized text extension
        if extension in OnyxFileExtensions.PLAIN_TEXT_EXTENSIONS:
            return extract_result_from_text_file(file)

        # If it's an image file or something else, we do not parse embedded images from them
        # just return empty text
        return ExtractionResult(text_content="", embedded_images=[], metadata={})

    except Exception as e:
        logger.exception(f"Failed to extract text/images from {file_name}: {e}")
        return ExtractionResult(text_content="", embedded_images=[], metadata={})


def docx_to_txt_filename(file_path: str) -> str:
    return file_path.rsplit(".", 1)[0] + ".txt"
